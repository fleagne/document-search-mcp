import meilisearch
from pathlib import Path
import docx
import openpyxl
from PIL import Image
import pytesseract
import xml.etree.ElementTree as ET
from colorama import Fore, Back, Style, init

# colorama初期化
init(autoreset=True)

# Meilisearchクライアント（ローカル起動前提）
client = meilisearch.Client('http://127.0.0.1:7700')
INDEX_NAME = 'documents'


def extract_text(file_path):
    """各形式からテキスト抽出（位置情報付き）"""
    path = Path(file_path)
    ext = path.suffix.lower()
    
    try:
        if ext == '.docx':
            doc = docx.Document(file_path)
            lines = []
            for i, p in enumerate(doc.paragraphs, 1):
                if p.text.strip():
                    lines.append(f"[段落{i}] {p.text}")
            return '\n'.join(lines)
        
        elif ext in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            lines = []
            
            for sheet in wb.worksheets:
                sheet_name = sheet.title
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    for col_idx, cell in enumerate(row, 1):
                        if cell is not None and str(cell).strip():
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            lines.append(f"[{sheet_name}][{col_letter}{row_idx}] {cell}")
            
            return '\n'.join(lines)
        
        elif ext == '.png':
            img = Image.open(file_path)
            text = pytesseract.image_to_string(img, lang='jpn')
            return f"[画像全体]\n{text}" if text.strip() else ""
        
        elif ext == '.drawio':
            tree = ET.parse(file_path)
            root = tree.getroot()
            lines = []
            for idx, elem in enumerate(root.iter(), 1):
                value = elem.get('value', '').strip()
                if value:
                    lines.append(f"[要素{idx}] {value}")
            return '\n'.join(lines)
        
        else:
            return ""
    
    except Exception as e:
        print(f"Error extracting {file_path}: {e}")
        return ""


def index_files(directory):
    """指定ディレクトリ内のファイルをインデックス化"""
    # インデックス作成（存在しない場合）
    try:
        index = client.get_index(INDEX_NAME)
    except:
        index = client.create_index(INDEX_NAME, {'primaryKey': 'id'})
    
    # サポート形式
    extensions = ['.docx', '.xlsx', '.xls', '.png', '.drawio']
    dir_path = Path(directory)
    
    if not dir_path.exists():
        print(f"Error: Directory '{directory}' does not exist")
        return
    
    files = []
    for ext in extensions:
        files.extend(dir_path.rglob(f'*{ext}'))
    
    print(f"Found {len(files)} files in {directory}")
    
    if len(files) == 0:
        print("No supported files found. Supported formats: .docx, .xlsx, .xls, .png, .drawio")
        return
    
    documents = []
    for i, file_path in enumerate(files):
        print(f"Processing {i+1}/{len(files)}: {file_path.name}")
        text = extract_text(file_path)
        
        if text.strip():
            documents.append({
                'id': str(i),
                'path': str(file_path),
                'filename': file_path.name,
                'content': text[:50000]  # 最大50000文字に拡張
            })
        else:
            print(f"  -> Skipped (no text extracted)")
    
    if len(documents) == 0:
        print("No documents with text content found")
        return
    
    print(f"\nIndexing {len(documents)} documents...")
    index.add_documents(documents)
    print("Done!")


def search_documents(query, limit=10):
    """検索実行"""
    try:
        index = client.get_index(INDEX_NAME)
    except:
        print(f"Error: Index '{INDEX_NAME}' does not exist. Please run 'index' command first.")
        return
    
    results = index.search(query, {
        'limit': limit,
        'attributesToHighlight': ['content'],
        'highlightPreTag': '<<<',
        'highlightPostTag': '>>>'
    })
    
    print(f"\n{Fore.CYAN}=== Search Results for '{query}' ==={Style.RESET_ALL}\n")
    
    if len(results['hits']) == 0:
        print("No results found.")
        return
    
    for idx, hit in enumerate(results['hits'], 1):
        print(f"{Fore.GREEN}{idx}. 📄 {hit['filename']}{Style.RESET_ALL}")
        print(f"   {Fore.YELLOW}Path:{Style.RESET_ALL} {hit['path']}")
        
        # ハイライト部分を抽出して表示
        if '_formatted' in hit and 'content' in hit['_formatted']:
            highlighted = hit['_formatted']['content']
            # ハイライト部分を抽出
            lines = highlighted.split('\n')
            matched_lines = [line for line in lines if '<<<' in line and '>>>' in line]
            
            if matched_lines:
                print(f"   {Fore.YELLOW}Matches:{Style.RESET_ALL}")
                for line in matched_lines[:5]:  # 最大5件表示
                    # ハイライトタグを色付きに変換
                    colored_line = line.replace('<<<', f'{Back.YELLOW}{Fore.BLACK}').replace('>>>', f'{Style.RESET_ALL}')
                    print(f"     {colored_line[:200]}")
        else:
            # フォールバック: 通常のプレビュー
            content_lines = hit['content'].split('\n')
            for line in content_lines[:3]:
                if line.strip():
                    print(f"     {line[:150]}")
        
        print()


def main():
    """メイン処理"""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python main.py index <directory>  # ファイルをインデックス化")
        print("  python main.py search <query>     # 検索")
        sys.exit(1)
    
    command = sys.argv[1]
    
    if command == 'index':
        if len(sys.argv) < 3:
            print("Error: directory path required")
            sys.exit(1)
        directory = sys.argv[2]
        index_files(directory)
    
    elif command == 'search':
        if len(sys.argv) < 3:
            print("Error: search query required")
            sys.exit(1)
        query = ' '.join(sys.argv[2:])
        search_documents(query)
    
    else:
        print(f"Unknown command: {command}")
        sys.exit(1)


if __name__ == '__main__':
    main()